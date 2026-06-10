from __future__ import annotations

import json
import os
import re
import urllib.error
import urllib.request
from abc import ABC, abstractmethod
from copy import deepcopy
from datetime import datetime, timezone
from io import BytesIO
from pathlib import Path
from typing import Any

from pydantic import BaseModel, ConfigDict, Field, ValidationError

from .models import (
    ComponentModel,
    ImportBridgeParamsResponse,
    ProjectBridge,
    ScenarioInput,
    StructureModel,
    UpperStructureComponent,
    WorkSection,
)
from .scenario_data import apply_resource_max_quantity_defaults


ONTOLOGY_PATH = Path(__file__).resolve().parent / "ontology" / "bridge_structure_ontology.v1.json"
SUPPORTED_SUFFIXES = {".xlsx", ".xlsm"}
DEFAULT_HEADER_DEPTH = 3


class BridgeImportError(ValueError):
    pass


class BridgeImportConfigError(RuntimeError):
    pass


class CanonicalBridgeDocument(BaseModel):
    model_config = ConfigDict(extra="allow")

    schemaVersion: str
    source: dict[str, Any]
    bridge: dict[str, Any]
    carriageways: list[dict[str, Any]] = Field(default_factory=list)
    referenceCatalogs: dict[str, Any] = Field(default_factory=dict)
    quality: dict[str, list[dict[str, Any]]] = Field(default_factory=lambda: {"checks": [], "warnings": []})


class BridgeImportAdapter(ABC):
    @abstractmethod
    def understand(self, workbook_facts: dict[str, Any], ontology: dict[str, Any]) -> dict[str, Any]:
        """Return Canonical Bridge JSON from normalized workbook facts."""


class LocalOntologyBridgeImportAdapter(BridgeImportAdapter):
    def understand(self, workbook_facts: dict[str, Any], ontology: dict[str, Any]) -> dict[str, Any]:
        return build_canonical_bridge(workbook_facts, ontology)


class HttpBridgeImportAdapter(BridgeImportAdapter):
    def __init__(self, endpoint: str, model: str | None, api_key: str | None) -> None:
        self.endpoint = endpoint
        self.model = model
        self.api_key = api_key

    def understand(self, workbook_facts: dict[str, Any], ontology: dict[str, Any]) -> dict[str, Any]:
        payload = json.dumps(
            {"model": self.model, "workbook_facts": workbook_facts, "ontology": ontology},
            ensure_ascii=False,
        ).encode("utf-8")
        headers = {"Content-Type": "application/json"}
        if self.api_key:
            headers["Authorization"] = f"Bearer {self.api_key}"
        request = urllib.request.Request(self.endpoint, data=payload, headers=headers, method="POST")
        try:
            with urllib.request.urlopen(request, timeout=60) as response:
                return json.loads(response.read().decode("utf-8"))
        except urllib.error.URLError as exc:
            raise BridgeImportConfigError(f"AI 适配器调用失败: {exc}") from exc


def get_bridge_import_adapter() -> BridgeImportAdapter:
    provider = os.getenv("BRIDGE_IMPORT_LLM_PROVIDER", "local").strip().lower()
    if provider in {"", "local", "heuristic", "fake"}:
        return LocalOntologyBridgeImportAdapter()
    if provider in {"http", "generic_http"}:
        endpoint = os.getenv("BRIDGE_IMPORT_LLM_ENDPOINT")
        if not endpoint:
            raise BridgeImportConfigError("BRIDGE_IMPORT_LLM_ENDPOINT 未配置，无法调用外部 AI 适配器。")
        return HttpBridgeImportAdapter(
            endpoint=endpoint,
            model=os.getenv("BRIDGE_IMPORT_LLM_MODEL"),
            api_key=os.getenv("BRIDGE_IMPORT_LLM_API_KEY"),
        )
    raise BridgeImportConfigError(f"不支持的桥梁导入 AI 适配器: {provider}")


def import_bridge_parameters(
    *,
    file_name: str,
    content: bytes,
    scenario: ScenarioInput,
    target_bridge: str | None = None,
    adapter: BridgeImportAdapter | None = None,
) -> ImportBridgeParamsResponse:
    ontology = load_bridge_ontology()
    workbook_facts = extract_workbook_facts(file_name=file_name, content=content, target_bridge=target_bridge, ontology=ontology)
    canonical_payload = (adapter or get_bridge_import_adapter()).understand(workbook_facts, ontology)
    try:
        canonical = CanonicalBridgeDocument.model_validate(canonical_payload)
    except ValidationError as exc:
        raise BridgeImportError(f"AI 结构理解结果不符合 Canonical Bridge JSON: {exc}") from exc

    next_scenario = scenario.model_copy(deep=True)
    next_scenario.project.bridges = canonical_to_project_bridges(canonical.model_dump(), ontology)
    apply_resource_max_quantity_defaults(next_scenario)
    summary = build_import_summary(canonical.model_dump())
    return ImportBridgeParamsResponse(
        scenario=next_scenario,
        canonical_bridge=canonical.model_dump(),
        summary=summary,
        quality_checks=canonical.quality.get("checks", []),
        warnings=canonical.quality.get("warnings", []),
    )


def load_bridge_ontology() -> dict[str, Any]:
    with ONTOLOGY_PATH.open("r", encoding="utf-8") as handle:
        return json.load(handle)


def extract_workbook_facts(
    *,
    file_name: str,
    content: bytes,
    target_bridge: str | None,
    ontology: dict[str, Any],
) -> dict[str, Any]:
    suffix = Path(file_name).suffix.lower()
    if Path(file_name).name.startswith("~$"):
        raise BridgeImportError("不能导入 Excel 临时锁文件，请关闭工作簿后选择正式文件。")
    if suffix not in SUPPORTED_SUFFIXES:
        raise BridgeImportError(f"暂不支持 {suffix or '无扩展名'} 文件，v1 仅支持 .xlsx/.xlsm。")
    if not content:
        raise BridgeImportError("上传的 Excel 文件为空。")

    try:
        workbook = _load_workbook(content)
    except BridgeImportError:
        raise
    except Exception as exc:
        raise BridgeImportError(f"Excel 工作簿读取失败: {exc}") from exc

    sheets: list[dict[str, Any]] = []
    for worksheet in workbook.worksheets:
        expanded = _expanded_sheet_values(worksheet)
        header_start = _find_header_start(expanded, worksheet.max_row, worksheet.max_column)
        if header_start is None:
            continue
        columns = _build_columns(expanded, header_start, worksheet.max_column)
        rows, sheet_warnings = _build_fact_rows(
            expanded=expanded,
            worksheet_name=worksheet.title,
            header_start=header_start,
            max_row=worksheet.max_row,
            columns=columns,
            ontology=ontology,
        )
        if rows:
            sheets.append(
                {
                    "name": worksheet.title,
                    "headerRows": list(range(header_start, header_start + DEFAULT_HEADER_DEPTH)),
                    "dataStartRow": header_start + DEFAULT_HEADER_DEPTH,
                    "columns": columns,
                    "rows": rows,
                    "warnings": sheet_warnings,
                }
            )

    if not sheets:
        raise BridgeImportError("未在工作簿中识别到包含桥梁结构参数的工作表。")

    return {
        "source": {
            "fileName": file_name,
            "targetBridge": target_bridge,
            "parsedAt": datetime.now(timezone.utc).isoformat(),
        },
        "sheets": sheets,
    }


def build_canonical_bridge(workbook_facts: dict[str, Any], ontology: dict[str, Any]) -> dict[str, Any]:
    rows = _matching_rows(workbook_facts)
    source = workbook_facts["source"]
    target_bridge = source.get("targetBridge")
    target_base = _normalize_bridge_base(str(target_bridge)) if target_bridge else _normalize_bridge_base(rows[0]["bridgeName"])
    selected = [row for row in rows if _row_matches_bridge(row, target_base)]
    if not selected:
        raise BridgeImportError(f"未找到目标桥梁 {target_bridge or target_base} 对应的结构参数行。")

    quality = _initial_quality(workbook_facts, selected)
    grouped: dict[str, list[dict[str, Any]]] = {}
    for row in selected:
        grouped.setdefault(row["side"], []).append(row)

    carriageways = []
    for side in _ordered_sides(grouped):
        side_rows = sorted(grouped[side], key=lambda item: (item.get("supportIndex") is None, item.get("supportIndex") or 0, item["rowNumber"]))
        span_groups = _build_span_groups(side_rows)
        supports = [_support_from_row(row, ontology, quality) for row in side_rows]
        _append_support_count_check(side, supports, span_groups, quality)
        continuous_blocks = [
            {
                "groupIndex": group["groupIndex"],
                "expression": group["expression"],
                "structureType": group["structureType"],
                "structureCode": group["structureCode"],
                "requiresSupplement": True,
                "source": "upperStructureAtSupport",
            }
            for group in span_groups
            if group["structureCode"] == "castInPlaceContinuousBoxGirder"
        ]
        if continuous_blocks:
            quality["warnings"].append(
                {
                    "id": f"continuous_beam_blocks_missing_{side}",
                    "message": f"{_side_label(side)}存在现浇/连续梁，但 Excel 中未提供分块或节段明细。",
                    "details": {"side": side, "blocks": continuous_blocks},
                }
            )
        carriageways.append(
            {
                "side": side,
                "name": side_rows[0]["bridgeName"],
                "centerStake": side_rows[0].get("centerStake"),
                "deck": _default_deck(side),
                "spanGroups": span_groups,
                "continuousBeamBlocks": continuous_blocks,
                "supports": supports,
            }
        )

    first_row = selected[0]
    first_spans = carriageways[0]["spanGroups"] if carriageways else []
    bridge = {
        "name": target_base,
        "centerStake": first_row.get("centerStake"),
        "startStake": None,
        "endStake": None,
        "lengthM": None,
        "totalWidthM": None,
        "skewAngleDeg": None,
        "spanExpression": "+".join(
            f"({group['expression']})" if group.get("structureCode") == "castInPlaceContinuousBoxGirder" else group["expression"]
            for group in first_spans
        ) or None,
        "upperStructureText": _join_unique(
            support.get("upperStructureAtSupport", {}).get("type")
            for carriageway in carriageways
            for support in carriageway["supports"]
        ),
        "pierFoundationText": "桩基础",
        "abutmentFoundationText": _join_unique(
            support.get("components", {}).get("abutment", {}).get("type")
            for carriageway in carriageways
            for support in carriageway["supports"]
        ),
        "remarks": None,
        "sourceTrace": {
            "sheet": first_row["sheetName"],
            "row": first_row["rowNumber"],
            "cells": first_row.get("sourceCells", {}),
        },
    }

    return {
        "schemaVersion": "bridge-parametric-excel/v1",
        "source": {
            **source,
            "targetBridge": target_bridge or target_base,
            "structureSheet": first_row["sheetName"],
        },
        "bridge": bridge,
        "carriageways": carriageways,
        "referenceCatalogs": {
            "ontologySchemaVersion": ontology.get("schema_version"),
            "adapter": os.getenv("BRIDGE_IMPORT_LLM_PROVIDER", "local"),
        },
        "quality": quality,
    }


def canonical_to_project_bridges(canonical: dict[str, Any], ontology: dict[str, Any]) -> list[ProjectBridge]:
    bridge_info = canonical["bridge"]
    source = canonical["source"]
    bridge = ProjectBridge(
        id="B1",
        name=bridge_info.get("name") or source.get("targetBridge") or "导入桥梁",
        order=1,
        workpoint_type="bridge",
        import_source={
            "fileName": source.get("fileName"),
            "parsedAt": source.get("parsedAt"),
            "schemaVersion": canonical.get("schemaVersion"),
            "ontologySchemaVersion": canonical.get("referenceCatalogs", {}).get("ontologySchemaVersion"),
        },
        work_sections=[],
    )

    for section_index, carriageway in enumerate(canonical.get("carriageways", []), start=1):
        side = carriageway.get("side") or "none"
        section = WorkSection(
            id=f"WS-{side.upper()}",
            name=f"{_side_label(side)}结构参数",
            order=section_index,
            side=side,
            structures=[],
            upper_structures=_upper_structures_from_span_groups(bridge.id, side, carriageway.get("spanGroups", [])),
        )
        for support in carriageway.get("supports", []):
            structure = _structure_from_support(support, bridge.id, side, ontology)
            section.structures.append(structure)
        bridge.work_sections.append(section)

    return [bridge]


def build_import_summary(canonical: dict[str, Any]) -> dict[str, Any]:
    carriageways = canonical.get("carriageways", [])
    supports = [support for carriageway in carriageways for support in carriageway.get("supports", [])]
    components = [
        component
        for support in supports
        for component in support.get("components", {}).values()
        if isinstance(component, dict) and component.get("present")
    ]
    upper_count = sum(group.get("totalSpanCount", 0) for carriageway in carriageways for group in carriageway.get("spanGroups", []))
    return {
        "bridgeName": canonical.get("bridge", {}).get("name"),
        "fileName": canonical.get("source", {}).get("fileName"),
        "carriagewayCount": len(carriageways),
        "supportCount": len(supports),
        "componentCount": len(components) + upper_count,
        "lowerComponentCount": len(components),
        "upperComponentCount": upper_count,
        "warningCount": len(canonical.get("quality", {}).get("warnings", [])),
        "sides": [carriageway.get("side") for carriageway in carriageways],
        "spanExpression": canonical.get("bridge", {}).get("spanExpression"),
    }


def _upper_structures_from_span_groups(
    bridge_id: str,
    side: str,
    span_groups: list[dict[str, Any]],
) -> list[UpperStructureComponent]:
    side_code = {"left": "L", "right": "R", "none": "N"}.get(side, "N")
    upper_structures: list[UpperStructureComponent] = []
    for group in span_groups:
        start_index = int(group.get("spanStartIndex") or 0)
        lengths = group.get("spanLengthsM") or []
        for offset, length in enumerate(lengths):
            span_index = start_index + offset
            support_range = f"{span_index - 1}#墩~{span_index}#墩"
            if span_index == 1:
                support_range = f"0#台~{span_index}#墩"
            if span_index == group.get("spanEndIndex"):
                total_spans = sum(item.get("totalSpanCount", 0) for item in span_groups)
                if span_index == total_spans:
                    support_range = f"{span_index - 1}#墩~{span_index}#台"
            structure_type = str(group.get("structureType") or "上部结构")
            upper_structures.append(
                UpperStructureComponent(
                    id=f"{bridge_id}-{side_code}-SPAN-{span_index:02d}",
                    name=f"{support_range}-{structure_type}",
                    structure_type=structure_type,
                    side=side,
                    span_index=span_index,
                    support_range=support_range,
                    span_length_m=float(length),
                    beam_count_per_span=group.get("beamCountPerSpan"),
                    span_group_expression=str(group.get("expression") or ""),
                    properties={
                        "structure_code": group.get("structureCode"),
                        "group_index": group.get("groupIndex"),
                        "span_count_per_unit": group.get("spanCountPerUnit"),
                        "same_type_span_count": group.get("sameTypeSpanCount"),
                        "total_span_count": group.get("totalSpanCount"),
                        "source": group.get("source"),
                    },
                )
            )
    return upper_structures


def _expanded_sheet_values(worksheet: Any) -> dict[tuple[int, int], Any]:
    values: dict[tuple[int, int], Any] = {}
    for row in range(1, worksheet.max_row + 1):
        for col in range(1, worksheet.max_column + 1):
            values[(row, col)] = worksheet.cell(row, col).value

    for merged_range in worksheet.merged_cells.ranges:
        top_value = worksheet.cell(merged_range.min_row, merged_range.min_col).value
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                values[(row, col)] = top_value
    return values


def _find_header_start(expanded: dict[tuple[int, int], Any], max_row: int, max_col: int) -> int | None:
    search_limit = min(max_row, 12)
    best: tuple[int, int] | None = None
    for row in range(1, search_limit + 1):
        text = " ".join(str(expanded.get((r, c)) or "") for r in range(row, min(row + DEFAULT_HEADER_DEPTH, max_row + 1)) for c in range(1, max_col + 1))
        score = sum(token in text for token in ["编号", "地名或桥名", "桥名", "下部结构", "桩基", "墩柱", "上部结构"])
        if "编号" in text and ("桩基" in text or "墩柱" in text or "桥台" in text):
            if best is None or score > best[1]:
                best = (row, score)
    return best[0] if best else None


def _build_columns(expanded: dict[tuple[int, int], Any], header_start: int, max_col: int) -> list[dict[str, Any]]:
    columns: list[dict[str, Any]] = []
    for col in range(1, max_col + 1):
        parts: list[str] = []
        for row in range(header_start, header_start + DEFAULT_HEADER_DEPTH):
            value = _clean_header(expanded.get((row, col)))
            if value and value not in parts:
                parts.append(value)
        header_path = " / ".join(parts)
        if not header_path:
            continue
        columns.append(
            {
                "index": col,
                "letter": _column_letter(col),
                "headerPath": header_path,
                "semanticKey": _semantic_key(header_path),
            }
        )
    return columns


def _build_fact_rows(
    *,
    expanded: dict[tuple[int, int], Any],
    worksheet_name: str,
    header_start: int,
    max_row: int,
    columns: list[dict[str, Any]],
    ontology: dict[str, Any],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    rows: list[dict[str, Any]] = []
    warnings: list[dict[str, Any]] = []
    last_bridge_name: str | None = None
    last_center_stake: str | None = None
    null_values = set(ontology.get("null_values", []))

    for row_number in range(header_start + DEFAULT_HEADER_DEPTH, max_row + 1):
        values: dict[str, Any] = {}
        raw: dict[str, Any] = {}
        source_cells: dict[str, str] = {}
        formula_errors = _formula_error_cells(expanded, row_number)

        for column in columns:
            key = column["semanticKey"]
            if not key:
                continue
            cell_value = expanded.get((row_number, column["index"]))
            normalized = _normalize_cell(cell_value, null_values)
            raw[key] = cell_value
            values[key] = normalized
            source_cells[key] = f"{column['letter']}{row_number}"

        bridge_name = values.get("bridge_name") or last_bridge_name
        center_stake = values.get("center_stake") or last_center_stake
        if values.get("bridge_name"):
            last_bridge_name = str(values["bridge_name"])
        if values.get("center_stake"):
            last_center_stake = str(values["center_stake"])

        support_no = values.get("support_no")
        if not support_no or not _parse_support_no(str(support_no)):
            continue
        if not bridge_name:
            warnings.append({"id": "missing_bridge_name", "message": f"{worksheet_name} 第 {row_number} 行缺少桥名，已跳过。"})
            continue

        support_index, support_type = _parse_support_no(str(support_no)) or (None, "unknown")
        if formula_errors:
            warnings.append(
                {
                    "id": "formula_error_normalized",
                    "message": f"{worksheet_name} 第 {row_number} 行存在公式错误，已按空值处理。",
                    "details": {"row": row_number, "cells": formula_errors},
                }
            )
        rows.append(
            {
                "sheetName": worksheet_name,
                "rowNumber": row_number,
                "bridgeName": str(bridge_name),
                "baseBridgeName": _normalize_bridge_base(str(bridge_name)),
                "side": _infer_side(str(bridge_name)),
                "centerStake": center_stake,
                "supportNo": str(support_no),
                "supportIndex": support_index,
                "supportType": support_type,
                "values": values,
                "raw": raw,
                "sourceCells": source_cells,
                "formulaErrors": formula_errors,
            }
        )
    return rows, warnings


def _matching_rows(workbook_facts: dict[str, Any]) -> list[dict[str, Any]]:
    rows = [row for sheet in workbook_facts.get("sheets", []) for row in sheet.get("rows", [])]
    if not rows:
        raise BridgeImportError("工作簿中未识别到任何墩台结构参数行。")
    return rows


def _row_matches_bridge(row: dict[str, Any], target_base: str) -> bool:
    row_base = _normalize_bridge_base(row.get("bridgeName", ""))
    return row_base == target_base or target_base in row_base or row_base in target_base


def _initial_quality(workbook_facts: dict[str, Any], selected_rows: list[dict[str, Any]]) -> dict[str, list[dict[str, Any]]]:
    sheet_names = [sheet["name"] for sheet in workbook_facts.get("sheets", [])]
    warnings = [warning for sheet in workbook_facts.get("sheets", []) for warning in sheet.get("warnings", [])]
    return {
        "checks": [
            {
                "id": "workbook_preprocessed",
                "status": "passed",
                "message": f"已识别 {len(sheet_names)} 个结构参数工作表，抽取 {len(selected_rows)} 行墩台数据。",
                "details": {"sheets": sheet_names},
            }
        ],
        "warnings": warnings,
    }


def _support_from_row(row: dict[str, Any], ontology: dict[str, Any], quality: dict[str, list[dict[str, Any]]]) -> dict[str, Any]:
    values = row["values"]
    components = {
        "pileFoundation": _pile_component(row),
        "spreadFoundation": {"present": False},
        "cap": _dimensioned_component(row, "cap", "承台"),
        "groundTieBeam": _dimensioned_component(row, "ground_tie", "地系梁"),
        "middleTieBeam": _dimensioned_component(row, "middle_tie", "中系梁"),
        "pierColumn": _pier_column_component(row),
        "capBeam": _dimensioned_component(row, "cap_beam", "盖梁"),
        "abutment": _abutment_component(row),
    }
    _append_pile_length_check(row, components["pileFoundation"], quality)
    foundation_type = _foundation_label("pileFoundation", ontology) if components["pileFoundation"]["present"] else None
    abutment_text = values.get("abutment_type")
    matched_foundation = _match_foundation_type(abutment_text, ontology)
    if matched_foundation:
        foundation_type = matched_foundation
    if foundation_type == _foundation_label("spreadFoundation", ontology):
        components["spreadFoundation"] = _spread_foundation_component(row)
    return {
        "supportNo": row["supportNo"],
        "supportIndex": row["supportIndex"],
        "supportType": row["supportType"],
        "foundationType": foundation_type or "unknown",
        "matchedModelHint": _matched_model_hint(row, components),
        "components": components,
        "upperStructureAtSupport": {
            "type": values.get("upper_structure_type"),
            "lengthM": _as_float(values.get("upper_length")),
            "count": _as_int(values.get("upper_quantity")),
            "sourceCells": _source_subset(row, ["upper_structure_type", "upper_length", "upper_quantity"]),
        },
        "sourceTrace": {
            "sheet": row["sheetName"],
            "row": row["rowNumber"],
            "bridgeName": row["bridgeName"],
            "cells": row.get("sourceCells", {}),
        },
    }


def _pile_component(row: dict[str, Any]) -> dict[str, Any]:
    values = row["values"]
    diameter_m = _cm_to_m(values.get("pile_diameter"))
    length_m = _cm_to_m(values.get("pile_length"))
    count = _as_int(values.get("pile_count"))
    total_length_m = _as_float(values.get("pile_total_length"))
    present = any(item is not None for item in [diameter_m, length_m, count, total_length_m])
    return {
        "present": present,
        "type": "桩基础" if present else None,
        "diameterM": diameter_m,
        "lengthM": length_m,
        "count": count,
        "totalLengthM": total_length_m,
        "raw": _raw_subset(row, ["pile_diameter", "pile_length", "pile_count", "pile_total_length"]),
        "sourceCells": _source_subset(row, ["pile_diameter", "pile_length", "pile_count", "pile_total_length"]),
    }


def _dimensioned_component(row: dict[str, Any], prefix: str, label: str) -> dict[str, Any]:
    dimensions_key = f"{prefix}_dimensions"
    count_key = f"{prefix}_count"
    values = row["values"]
    dimensions = _parse_dimensions_m(values.get(dimensions_key))
    count = _as_int(values.get(count_key))
    present = dimensions is not None or count is not None
    return {
        "present": present,
        "type": label if present else None,
        "dimensionsM": dimensions,
        "count": count,
        "raw": _raw_subset(row, [dimensions_key, count_key]),
        "sourceCells": _source_subset(row, [dimensions_key, count_key]),
    }


def _pier_column_component(row: dict[str, Any]) -> dict[str, Any]:
    values = row["values"]
    dimensions = _parse_dimensions_m(values.get("pier_dimensions"))
    height_m = _cm_to_m(values.get("pier_height"))
    count = _as_int(values.get("pier_count"))
    form = values.get("pier_form")
    present = any(item is not None for item in [dimensions, height_m, count, form])
    return {
        "present": present,
        "type": form or ("墩柱" if present else None),
        "dimensionsM": dimensions,
        "heightM": height_m,
        "count": count,
        "raw": _raw_subset(row, ["pier_form", "pier_dimensions", "pier_height", "pier_count"]),
        "sourceCells": _source_subset(row, ["pier_form", "pier_dimensions", "pier_height", "pier_count"]),
    }


def _abutment_component(row: dict[str, Any]) -> dict[str, Any]:
    abutment_type = row["values"].get("abutment_type")
    present = abutment_type is not None
    return {
        "present": present,
        "type": abutment_type,
        "count": 1 if present else None,
        "raw": _raw_subset(row, ["abutment_type"]),
        "sourceCells": _source_subset(row, ["abutment_type"]),
    }


def _spread_foundation_component(row: dict[str, Any]) -> dict[str, Any]:
    return {
        "present": True,
        "type": "扩大基础",
        "count": 1,
        "raw": _raw_subset(row, ["abutment_type"]),
        "sourceCells": _source_subset(row, ["abutment_type"]),
    }


def _build_span_groups(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    spans: list[dict[str, Any]] = []
    for row in rows:
        length = _as_float(row["values"].get("upper_length"))
        structure_type = row["values"].get("upper_structure_type")
        count = _as_int(row["values"].get("upper_quantity"))
        if length is None or not structure_type:
            continue
        spans.append(
            {
                "length": length,
                "type": str(structure_type),
                "count": count,
                "row": row,
                "code": _upper_structure_code(str(structure_type)),
            }
        )
    groups: list[dict[str, Any]] = []
    current: list[dict[str, Any]] = []
    for span in spans:
        if not current:
            current = [span]
            continue
        previous = current[-1]
        same_continuous = span["code"] == previous["code"] == "castInPlaceContinuousBoxGirder"
        same_simple = span["code"] == previous["code"] and span["length"] == previous["length"]
        if same_continuous or same_simple:
            current.append(span)
        else:
            groups.append(_span_group_payload(current, len(groups) + 1))
            current = [span]
    if current:
        groups.append(_span_group_payload(current, len(groups) + 1))
    cursor = 1
    for group in groups:
        group["spanStartIndex"] = cursor
        cursor += group["totalSpanCount"]
        group["spanEndIndex"] = cursor - 1
    return groups


def _span_group_payload(spans: list[dict[str, Any]], index: int) -> dict[str, Any]:
    code = spans[0]["code"]
    lengths = [span["length"] for span in spans]
    if code == "castInPlaceContinuousBoxGirder":
        expression = "+".join(_format_number(length) for length in lengths)
        same_type_count = 1
        span_count_per_unit = len(spans)
    else:
        expression = f"{len(spans)}*{_format_number(lengths[0])}" if len(spans) > 1 else _format_number(lengths[0])
        same_type_count = len(spans)
        span_count_per_unit = 1
    return {
        "groupIndex": index,
        "expression": expression,
        "spanCountPerUnit": span_count_per_unit,
        "sameTypeSpanCount": same_type_count,
        "totalSpanCount": len(spans),
        "spanLengthsM": lengths,
        "spanStartIndex": None,
        "spanEndIndex": None,
        "structureType": "现浇连续梁" if code == "castInPlaceContinuousBoxGirder" else "简支T梁",
        "structureCode": code,
        "beamHeightM": None,
        "beamCountPerSpan": spans[0].get("count"),
        "deckWidthM": None,
        "horizontalOffsetM": None,
        "source": "upperStructureAtSupport",
    }


def _append_support_count_check(
    side: str,
    supports: list[dict[str, Any]],
    span_groups: list[dict[str, Any]],
    quality: dict[str, list[dict[str, Any]]],
) -> None:
    span_count = sum(group.get("totalSpanCount", 0) for group in span_groups)
    if not span_count:
        quality["checks"].append(
            {"id": f"support_count_{side}", "status": "skipped", "message": f"{_side_label(side)}未识别到跨径数量，跳过墩台数量核对。"}
        )
        return
    status = "passed" if len(supports) == span_count + 1 else "failed"
    quality["checks"].append(
        {
            "id": f"support_count_{side}",
            "status": status,
            "message": f"{_side_label(side)}墩台数量 {len(supports)}，跨数 {span_count}。",
            "details": {"side": side, "supportCount": len(supports), "spanCount": span_count},
        }
    )


def _append_pile_length_check(row: dict[str, Any], pile: dict[str, Any], quality: dict[str, list[dict[str, Any]]]) -> None:
    if not pile.get("present"):
        return
    length = pile.get("lengthM")
    count = pile.get("count")
    total = pile.get("totalLengthM")
    if length is None or count is None or total is None:
        quality["checks"].append(
            {
                "id": f"pile_total_{row['side']}_{row['supportIndex']}",
                "status": "skipped",
                "message": f"{row['bridgeName']} {row['supportNo']} 桩长/根数/总长信息不完整，跳过核对。",
            }
        )
        return
    expected = length * count
    status = "passed" if abs(expected - total) <= 0.05 else "failed"
    quality["checks"].append(
        {
            "id": f"pile_total_{row['side']}_{row['supportIndex']}",
            "status": status,
            "message": f"{row['bridgeName']} {row['supportNo']} 桩长 × 根数 = {_format_number(expected)}m，表内总长 {_format_number(total)}m。",
            "details": {
                "row": row["rowNumber"],
                "expectedTotalLengthM": expected,
                "actualTotalLengthM": total,
                "sourceCells": pile.get("sourceCells", {}),
            },
        }
    )


def _structure_from_support(support: dict[str, Any], bridge_id: str, side: str, ontology: dict[str, Any]) -> StructureModel:
    support_index = support.get("supportIndex")
    support_type = support.get("supportType") if support.get("supportType") in {"pier", "abutment"} else "pier"
    side_code = {"left": "L", "right": "R", "none": "N"}.get(side, "N")
    structure_code = "A" if support_type == "abutment" else "P"
    structure_id = f"{bridge_id}-{side_code}-{structure_code}{int(support_index or 0):02d}"
    structure = StructureModel(
        id=structure_id,
        name=support.get("supportNo") or structure_id,
        structure_type=support_type,
        order=int(support_index or 0),
        support_no=support.get("supportNo"),
        support_index=support_index,
        components=[],
    )
    components = support.get("components", {})
    structure.components.extend(_pile_components_for_schedule(structure_id, structure.name, components.get("pileFoundation", {}), support))
    _append_component_if_present(
        structure.components,
        structure_id,
        structure.name,
        "SPREAD-FOUNDATION",
        "扩大基础",
        "spread_foundation",
        components.get("spreadFoundation", {}),
        support,
        quantity=_count_quantity(components.get("spreadFoundation", {})),
        quantity_label=_dimensions_label(components.get("spreadFoundation", {})),
    )
    _append_component_if_present(
        structure.components,
        structure_id,
        structure.name,
        "CAP",
        "承台",
        "cap",
        components.get("cap", {}),
        support,
        quantity=_count_quantity(components.get("cap", {})),
        quantity_label=_dimensions_label(components.get("cap", {})),
    )
    _append_component_if_present(
        structure.components,
        structure_id,
        structure.name,
        "GROUND-TIE",
        "地系梁",
        "ground_tie_beam",
        components.get("groundTieBeam", {}),
        support,
        quantity=_count_quantity(components.get("groundTieBeam", {})),
        quantity_label=_dimensions_label(components.get("groundTieBeam", {})),
    )
    pier_column = components.get("pierColumn", {})
    if pier_column.get("present"):
        height = pier_column.get("heightM")
        count = pier_column.get("count") or 1
        quantity = (height or 0) * count if height else count
        _append_component_if_present(
            structure.components,
            structure_id,
            structure.name,
            "BODY",
            str(pier_column.get("type") or "墩柱"),
            "pier_body",
            pier_column,
            support,
            quantity=quantity,
            quantity_label=f"{count}根，高度{_format_number(height)}m" if height else f"{count}根",
        )
    _append_component_if_present(
        structure.components,
        structure_id,
        structure.name,
        "MIDDLE-TIE",
        "中系梁",
        "middle_tie_beam",
        components.get("middleTieBeam", {}),
        support,
        quantity=_count_quantity(components.get("middleTieBeam", {})),
        quantity_label=_dimensions_label(components.get("middleTieBeam", {})),
    )
    _append_component_if_present(
        structure.components,
        structure_id,
        structure.name,
        "BEAM",
        "盖梁",
        "cap_beam",
        components.get("capBeam", {}),
        support,
        quantity=_count_quantity(components.get("capBeam", {})),
        quantity_label=_dimensions_label(components.get("capBeam", {})),
    )
    _append_component_if_present(
        structure.components,
        structure_id,
        structure.name,
        "ABUTMENT",
        str(components.get("abutment", {}).get("type") or "桥台"),
        "abutment_body",
        components.get("abutment", {}),
        support,
        quantity=1,
        quantity_label=str(components.get("abutment", {}).get("type") or "1个"),
    )
    return structure


def _pile_components_for_schedule(
    structure_id: str,
    structure_name: str,
    pile: dict[str, Any],
    support: dict[str, Any],
) -> list[ComponentModel]:
    if not pile.get("present") or not pile.get("lengthM"):
        return []
    count = pile.get("count") or 1
    components = []
    for pile_no in range(1, int(count) + 1):
        components.append(
            ComponentModel(
                id=f"{structure_id}-PILE-{pile_no:02d}",
                name=f"{structure_name}-{pile_no}#桩基",
                component_type="pile",
                quantity=float(pile["lengthM"]),
                quantity_label=_pile_label(pile.get("diameterM"), pile.get("lengthM")),
                method_id="rotary_drill",
                properties=_component_properties("pileFoundation", pile, support, {"pile_no": pile_no}),
            )
        )
    return components


def _append_component_if_present(
    target: list[ComponentModel],
    structure_id: str,
    structure_name: str,
    suffix: str,
    name_suffix: str,
    component_type: str,
    canonical_component: dict[str, Any],
    support: dict[str, Any],
    *,
    quantity: float,
    quantity_label: str,
) -> None:
    if not canonical_component.get("present") or quantity <= 0:
        return
    target.append(
        ComponentModel(
            id=f"{structure_id}-{suffix}",
            name=f"{structure_name}-{name_suffix}",
            component_type=component_type,
            quantity=quantity,
            quantity_label=quantity_label,
            properties=_component_properties(suffix.lower(), canonical_component, support),
        )
    )


def _component_properties(
    ontology_code: str,
    canonical_component: dict[str, Any],
    support: dict[str, Any],
    extra: dict[str, Any] | None = None,
) -> dict[str, Any]:
    dimensions = canonical_component.get("dimensionsM")
    if dimensions is None and canonical_component.get("diameterM") is not None:
        dimensions = {"diameterM": canonical_component.get("diameterM"), "lengthM": canonical_component.get("lengthM")}
    payload = {
        "ontology_code": ontology_code,
        "dimensions_m": dimensions,
        "raw": deepcopy(canonical_component.get("raw") or {}),
        "source_trace": {
            **support.get("sourceTrace", {}),
            "componentCells": canonical_component.get("sourceCells", {}),
        },
        "confidence": 0.82,
        "import_warnings": [],
    }
    if extra:
        payload.update(extra)
    return payload


def _count_quantity(component: dict[str, Any]) -> float:
    return float(component.get("count") or 1)


def _dimensions_label(component: dict[str, Any]) -> str:
    dimensions = component.get("dimensionsM")
    count = component.get("count") or 1
    if dimensions:
        return f"{_format_dimensions(dimensions)}，{count}个"
    return f"{count}个"


def _default_deck(side: str) -> dict[str, Any]:
    offsets = {"left": -8.5, "right": 8.5, "none": 0}
    return {"widthM": 12.6 if side in {"left", "right"} else None, "offsetM": offsets.get(side, 0), "source": "defaultFromOntology"}


def _semantic_key(header_path: str) -> str | None:
    compact = header_path.replace(" ", "")
    if "序号" in compact:
        return "serial_no"
    if "桥跨中心桩号" in compact or "中心桩号" in compact:
        return "center_stake"
    if "地名或桥名" in compact or ("桥名" in compact and "桥台" not in compact):
        return "bridge_name"
    if compact.endswith("编号") or compact == "编号":
        return "support_no"
    if "桩基" in compact and "桩径" in compact:
        return "pile_diameter"
    if "桩基" in compact and "桩长" in compact:
        return "pile_length"
    if "桩基" in compact and "根数" in compact:
        return "pile_count"
    if "桩基" in compact and "总长" in compact:
        return "pile_total_length"
    if "承台" in compact and "尺寸" in compact:
        return "cap_dimensions"
    if "承台" in compact and "数量" in compact:
        return "cap_count"
    if "地系梁" in compact and "尺寸" in compact:
        return "ground_tie_dimensions"
    if "地系梁" in compact and "数量" in compact:
        return "ground_tie_count"
    if "中系梁" in compact and "尺寸" in compact:
        return "middle_tie_dimensions"
    if "中系梁" in compact and "数量" in compact:
        return "middle_tie_count"
    if "墩柱" in compact and "形式" in compact:
        return "pier_form"
    if "墩柱" in compact and "尺寸" in compact:
        return "pier_dimensions"
    if "墩柱" in compact and "高度" in compact:
        return "pier_height"
    if "墩柱" in compact and "数量" in compact:
        return "pier_count"
    if "盖梁" in compact and "尺寸" in compact:
        return "cap_beam_dimensions"
    if "盖梁" in compact and "数量" in compact:
        return "cap_beam_count"
    if compact.endswith("桥台") or "下部结构/桥台" in compact:
        return "abutment_type"
    if "上部结构" in compact and "长度" in compact:
        return "upper_length"
    if "上部结构" in compact and ("数量" in compact or "片" in compact):
        return "upper_quantity"
    if "上部结构" in compact and ("类型" in compact or "结构形式" in compact):
        return "upper_structure_type"
    return None


def _clean_header(value: Any) -> str | None:
    if value is None:
        return None
    text = re.sub(r"\s+", "", str(value))
    return text or None


def _normalize_cell(value: Any, null_values: set[str]) -> Any:
    if value is None:
        return None
    if isinstance(value, str):
        text = value.strip()
        if text in null_values or text.startswith("#"):
            return None
        return text
    return value


def _formula_error_cells(expanded: dict[tuple[int, int], Any], row_number: int) -> list[str]:
    cells: list[str] = []
    for row, col in expanded:
        if row != row_number:
            continue
        value = expanded[(row, col)]
        if isinstance(value, str) and value.strip().startswith("#"):
            cells.append(f"{_column_letter(col)}{row_number}")
    return cells


def _load_workbook(content: bytes) -> Any:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise BridgeImportError("缺少 openpyxl 依赖，请先执行 pip install -r requirements.txt。") from exc
    return load_workbook(BytesIO(content), data_only=True, read_only=False)


def _column_letter(index: int) -> str:
    letters = ""
    while index:
        index, remainder = divmod(index - 1, 26)
        letters = chr(65 + remainder) + letters
    return letters


def _infer_side(bridge_name: str) -> str:
    if any(alias in bridge_name for alias in ["左幅", "左线", "左侧"]):
        return "left"
    if any(alias in bridge_name for alias in ["右幅", "右线", "右侧"]):
        return "right"
    return "none"


def _normalize_bridge_base(bridge_name: str) -> str:
    text = re.sub(r"\s+", "", str(bridge_name))
    return re.sub(r"(左幅|右幅|左线|右线|左侧|右侧|整幅|全幅)$", "", text)


def _parse_support_no(text: str) -> tuple[int | None, str] | None:
    match = re.search(r"(\d+)\s*#\s*([墩台])", text)
    if not match:
        return None
    return int(match.group(1)), "abutment" if match.group(2) == "台" else "pier"


def _parse_dimensions_m(value: Any) -> list[float] | None:
    if value is None:
        return None
    if isinstance(value, int | float):
        return [_round_m(float(value) / 100)]
    numbers = re.findall(r"\d+(?:\.\d+)?", str(value))
    if not numbers:
        return None
    return [_round_m(float(number) / 100) for number in numbers]


def _cm_to_m(value: Any) -> float | None:
    number = _as_float(value)
    return _round_m(number / 100) if number is not None else None


def _as_float(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, int | float):
        return float(value)
    match = re.search(r"-?\d+(?:\.\d+)?", str(value))
    return float(match.group(0)) if match else None


def _as_int(value: Any) -> int | None:
    number = _as_float(value)
    return int(number) if number is not None else None


def _round_m(value: float) -> float:
    return round(value, 4)


def _raw_subset(row: dict[str, Any], keys: list[str]) -> dict[str, Any]:
    return {key: row.get("raw", {}).get(key) for key in keys if row.get("raw", {}).get(key) is not None}


def _source_subset(row: dict[str, Any], keys: list[str]) -> dict[str, str]:
    return {key: row.get("sourceCells", {}).get(key) for key in keys if row.get("sourceCells", {}).get(key)}


def _matched_model_hint(row: dict[str, Any], components: dict[str, dict[str, Any]]) -> str:
    labels = [component.get("type") for component in components.values() if component.get("present") and component.get("type")]
    if not labels:
        return row["supportType"]
    return "+".join(str(label) for label in labels)


def _match_foundation_type(text: Any, ontology: dict[str, Any]) -> str | None:
    if text is None:
        return None
    value = str(text)
    for foundation in ontology.get("foundation_types", {}).values():
        aliases = foundation.get("aliases", [])
        if any(alias and alias in value for alias in aliases):
            return foundation.get("label")
    return None


def _foundation_label(code: str, ontology: dict[str, Any]) -> str:
    return ontology.get("foundation_types", {}).get(code, {}).get("label", code)


def _ordered_sides(grouped: dict[str, list[dict[str, Any]]]) -> list[str]:
    order = {"left": 0, "right": 1, "none": 2}
    return sorted(grouped, key=lambda side: order.get(side, 99))


def _side_label(side: str) -> str:
    return {"left": "左幅", "right": "右幅", "none": "无幅别"}.get(side, side)


def _upper_structure_code(text: str) -> str:
    if any(token in text for token in ["现浇", "连续", "刚构"]):
        return "castInPlaceContinuousBoxGirder"
    return "precastTGirder"


def _join_unique(values: Any) -> str | None:
    seen: list[str] = []
    for value in values:
        if value is None:
            continue
        text = str(value)
        if text and text not in seen:
            seen.append(text)
    return " / ".join(seen) if seen else None


def _format_number(value: Any) -> str:
    if value is None:
        return "-"
    number = float(value)
    return str(int(number)) if number.is_integer() else f"{number:g}"


def _pile_label(diameter_m: Any, length_m: Any) -> str:
    parts = []
    if diameter_m is not None:
        parts.append(f"直径{_format_number(diameter_m)}m")
    if length_m is not None:
        parts.append(f"桩长{_format_number(length_m)}m")
    return "，".join(parts) if parts else "-"


def _format_dimensions(dimensions: list[float]) -> str:
    return " × ".join(f"{_format_number(value)}m" for value in dimensions)
